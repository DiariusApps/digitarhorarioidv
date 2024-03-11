# pandas to xlsx python
from openpyxl import load_workbook
import io
day_extenso = {
    'SEG': 'SEGUNDA-FEIRA',
    'TER': 'TERÇA-FEIRA',
    'QUA': 'QUARTA-FEIRA',
    'QUI': 'QUINTA-FEIRA',
    'SEX': 'SEXTA-FEIRA',
    'SAB': 'SÁBADO',
    'DOM': 'DOMINGO'
}


def pandas_to_xlsx(df, day, turno='DIURNO'):
    if turno == 'DIURNO':
        wb = load_workbook(filename='hora.xlsx')
    else:
        wb = load_workbook(filename='noturno.xlsx')
    sheet_ranges = wb['hora']
    sheet_ranges['A1'] = f'{day_extenso[day]}'
    for j, row in enumerate(df.iterrows()):
        valores = row[1].tolist()
        index = row[0]
        sheet_ranges[f'A{j+3}'] = index.replace('1', '1°').replace('2', '2°').replace('3', '3°')
        for i, valor in enumerate(valores):
            sheet_ranges[f'{chr(65+i+1)}{j+3}'] = valor
    # save as stream
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()

def by_class(df, turma):
    wb = load_workbook(filename='HORARIO_por_turma.xlsx')
    sheet_ranges = wb['Planilha1']
    sheet_ranges['A2'] = turma
    for j, row in enumerate(df.iterrows()):
        valores = row[1].tolist()
        for i, valor in enumerate(valores):
            cell = sheet_ranges.cell(row=j+5, column=i+2)  # Using cell method to access individual cells
            cell.value = valor
            if 'empreendedorismo' in str(valor).lower():
                cell.font = cell.font.copy(size=12)
            elif 'VII_IDENTIDAD' in str(valor):
                cell.font = cell.font.copy(size=14)
            elif '_' in str(valor):
                cell.font = cell.font.copy(size=16)
            else:
                pass
    # save as stream
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()

def by_class_5h(df, turma):
    wb = load_workbook(filename='HORARIO_por_turma_5_horas.xlsx')
    sheet_ranges = wb['Planilha1']
    sheet_ranges['A2'] = turma
    for j, row in enumerate(df.iterrows()):
        valores = row[1].tolist()
        for i, valor in enumerate(valores):
            cell = sheet_ranges.cell(row=j+5, column=i+2)  # Using cell method to access individual cells
            cell.value = valor
            if len(str(valor)) > 36:
                cell.font = cell.font.copy(size=18)
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


